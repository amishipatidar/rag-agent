"""
Document Parsers - Extract text content from Word, Excel, and PDF files.
"""
import os
from typing import List


def parse_word(file_path: str) -> str:
    """
    Extract text from a .docx file.

    Args:
        file_path: Path to the .docx file.

    Returns:
        Extracted text as a single string, paragraphs separated by newlines.

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If the file is not a .docx file.
    """
    if not file_path.lower().endswith(".docx"):
        raise ValueError(f"Expected a .docx file, got: {file_path}")
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    from docx import Document
    doc = Document(file_path)
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n".join(paragraphs)


def parse_excel(file_path: str) -> str:
    """
    Extract text from an .xlsx file.
    Reads all sheets and converts each row into a comma-separated line.

    Args:
        file_path: Path to the .xlsx file.

    Returns:
        Extracted text as a single string.

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If the file is not an .xlsx file.
    """
    if not file_path.lower().endswith(".xlsx"):
        raise ValueError(f"Expected an .xlsx file, got: {file_path}")
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    from openpyxl import load_workbook
    wb = load_workbook(file_path, read_only=True, data_only=True)
    lines = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"--- Sheet: {sheet_name} ---")
        for row in ws.iter_rows(values_only=True):
            # Convert each cell to string, skip None values
            cells = [str(cell) if cell is not None else "" for cell in row]
            row_text = ", ".join(cells)
            if row_text.strip(", "):
                lines.append(row_text)

    wb.close()
    return "\n".join(lines)


def parse_pdf(file_path: str) -> str:
    """
    Extract text from a .pdf file.

    Args:
        file_path: Path to the .pdf file.

    Returns:
        Extracted text as a single string, pages separated by newlines.

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If the file is not a .pdf file.
    """
    if not file_path.lower().endswith(".pdf"):
        raise ValueError(f"Expected a .pdf file, got: {file_path}")
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    import pdfplumber
    text_parts = []

    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text_parts.append(page_text)

    return "\n".join(text_parts)


def parse_document(file_path: str) -> str:
    """
    Auto-detect file type and parse accordingly.

    Args:
        file_path: Path to the document file.

    Returns:
        Extracted text content.

    Raises:
        ValueError: If the file type is not supported.
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".docx":
        return parse_word(file_path)
    elif ext == ".xlsx":
        return parse_excel(file_path)
    elif ext == ".pdf":
        return parse_pdf(file_path)
    else:
        raise ValueError(
            f"Unsupported file type: '{ext}'. "
            f"Supported types: .docx, .xlsx, .pdf"
        )
